import xlwt
from django.shortcuts import render,redirect

# Create your views here.
from django.http import HttpResponse
import cv2  # openCV
import numpy as np  # for numpy arrays
#import sqlite3
import dlib
import os,time  # for creating folders
from .models import Student
import sys
import cognitive_face as CF
import urllib
from .forms import *
from openpyxl import Workbook, load_workbook
from .models import *
from .util import *
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

def index(request):
    return render(request,'home.html')
def imageForm(request):
    form = ImageForm()
    return render(request,'imageupload.html',{'form': form})

def signup(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            emailt=form.cleaned_data.get("emailt")
            exist = teacherEmailExist(emailt)
            if exist == False:
                name = form.cleaned_data.get("name")
                password = form.cleaned_data.get("password")
                obj = Signup.objects.create(Emailt=emailt, Name=name, Password=password)
                obj.save()
                request.session['emailt'] = emailt
                return render(request,'home.html')
            else:
                return HttpResponse('data already exists')
    else:
        form = SignupForm()
        print("error")
    return render(request, 'signup.html', {'form': form})
def login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            emailt=form.cleaned_data.get("emailt")
            exist = teacherEmailExist(emailt)
            if exist == True:
                password = form.cleaned_data.get("password")
                obj = Signup.objects.get(Emailt=emailt)
                if obj.Password ==password:
                    request.session['emailt']=emailt
                    return render(request,'home.html')
                else:
                    return HttpResponse('success')
            else:
                return HttpResponse('login Id is not exists')
    else:
        form = LoginForm()
        print("error")
    return render(request, 'login.html', {'form': form})

def addStudent(request):
    cap = cv2.VideoCapture(0)
    detector = dlib.get_frontal_face_detector()
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            emails=form.cleaned_data.get("emails")
            exist = studentEmailExist(emails)
            if exist == False:
                # print(name)
                # print(roll)
                roll = form.cleaned_data.get("roll")
                Id = roll[-2:]
                # addStudents(name, roll)                                                  # calling the sqlite3 database

                folderName = "user" + Id  # creating the person or user folder
                folderPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), "dataset/" + folderName)
                if not os.path.exists(folderPath):
                    os.makedirs(folderPath)

                sampleNum = 0
                while (True):
                    ret, img = cap.read()  # reading the camera input
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # Converting to GrayScale
                    dets = detector(img, 1)
                    for i, d in enumerate(dets):  # loop will run for each face detected
                        sampleNum += 1
                        cv2.imwrite(folderPath + "/User." + Id + "." + str(sampleNum) + ".jpg",
                                    img[d.top():d.bottom(), d.left():d.right()])  # Saving the faces
                        cv2.rectangle(img, (d.left(), d.top()), (d.right(), d.bottom()), (0, 255, 0),
                                      2)  # Forming the rectangle
                        cv2.waitKey(200)  # waiting time of 200 milisecond
                    cv2.imshow('frame', img)  # showing the video input from camera on window
                    cv2.waitKey(1)
                    if (sampleNum >= 20):  # will take 20 faces
                        break

                cap.release()  # turning the webcam off
                cv2.destroyAllWindows()

                emailt=request.session['emailt']
                name = form.cleaned_data.get("name")

                branch=form.cleaned_data.get("branch")
                sem = form.cleaned_data.get("sem")
                div=form.cleaned_data.get("div")
                obj = Student.objects.create(Emailt=emailt,Emails=emails, Name=name, Roll=roll,Branch=branch,Sem=sem,Div=div)
                obj.save()
                request.session['roll']=roll
                return redirect('creategroup')
            else:
                return HttpResponse('data already exists')
    else:
        form = StudentForm()
        print("error")
    return render(request, 'studentregistration.html', {'form': form})

def createGroup(request):

    Key = '97849dde33244e88a570f60327c27604'
    CF.Key.set(Key)
    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)
    roll=request.session['roll']
    if roll is not None:
        print(str('user'+roll[-2:]))
        res = CF.person.create(personGroupId(), str('user'+roll[-2:]))  # getting personid
        print(res)
        check=updatePersonId(roll,res)
        if check:
            print("Person ID successfully added to the database")
            #addPersonFaces(roll)
            #train()
        else:
            print("Record doesnt found")
            return redirect("error")

    return redirect('addpersonfaces',)

def addPersonFaces(request):

    Key = '97849dde33244e88a570f60327c27604'
    CF.Key.set(Key)

    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)
    roll=request.session.pop('roll')
    if roll is not None:
        currentDir = os.path.dirname(os.path.abspath(__file__))
        imageFolder = os.path.join(currentDir, "dataset/" + str('user'+roll[-2:]))
        person_id = getPersonIdFromDatabase(roll)
        for filename in os.listdir(imageFolder):
            if filename.endswith(".jpg"):
                print(filename)
                imgurl = os.path.join(imageFolder, filename)
                res = CF.face.detect(imgurl)
                if len(res) != 1:
                    print("No face detected in image")
                else:
                    res = CF.person.add_face(imgurl, personGroupId(), person_id)
                    print(res)
                time.sleep(6)
        train()
        return render(request,'home.html')
def train():

    Key = '97849dde33244e88a570f60327c27604'
    CF.Key.set(Key)

    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'  # Replace with your regional Base URL
    # https://westcentralus.api.cognitive.microsoft.com/face/v1.0/
    CF.BaseUrl.set(BASE_URL)

    res = CF.person_group.train(personGroupId())
    print(res)
    checkStatus()
def checkStatus():
    Key = '97849dde33244e88a570f60327c27604'
    CF.Key.set(Key)

    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)

    res = CF.person_group.get_status(personGroupId())
    print(res)
    if res['status'] != 'succeeded':
        return HttpResponse('model was not trained')
    # return render(request,'status.html')

def imageUpload(request):
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        # print(form.Meta.model.filename(self))
        if form.is_valid():
            img = form.cleaned_data.get("imageInput")
            emailt=request.session['emailt']
            subject=form.cleaned_data.get("subject")
            branch = form.cleaned_data.get("branch")
            sem = form.cleaned_data.get("sem")
            div = form.cleaned_data.get("div")
            obj=Image.objects.create(Imagepath=img,Emailt=emailt,Subject=subject,Branch=branch,Sem=sem,Div=div)
            imgName=obj.filename()
            obj.save()
            detectFace(imgName)

            getSpreadsheet(emailt=request.session['emailt'],subject=subject,branch=branch,sem=sem,div=div)
            identify(emailt=request.session['emailt'],subject=subject,branch=branch,sem=sem,div=div)
            return render(request,'home.html')
    else:
        form = ImageForm()
    return render(request, 'imageupload.html', {'form': form})

def detectFace(imgName):
    detector = dlib.get_frontal_face_detector()
    #if len
    img = cv2.imread('./media/images/'+imgName)
    dets = detector(img, 1)
    if not os.path.exists('./polls/Cropped_faces'):
        os.makedirs('./polls/Cropped_faces')
    print("detected = " + str(len(dets)))
    if len(dets)!=0:
        for i, d in enumerate(dets):
            cv2.imwrite('./polls/Cropped_faces/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])
        return
    else:
        return HttpResponse("face not detected")

def getSpreadsheet(emailt,subject,branch,sem,div):
    currentDate = time.strftime("%d_%m_%y")

    # create a workbook and add a worksheet
    reportName = emailt + "_" + subject + "_" + branch + "_" + sem + "_" + div + '.xlsx'
    if (os.path.exists('./'+reportName)):
        wb = load_workbook(filename=reportName)
        sheet = wb.get_sheet_by_name('Cse15')
        # sheet[ord() + '1']

        sheet_obj = wb.active
        m_row = sheet_obj.max_row
        m_column = sheet_obj.max_column
        d = {}

        for i in range(3, m_row + 1):
            key = sheet_obj.cell(row=i, column=1).value
            v = []
            for j in range(2, m_column + 1):
                v1 = sheet_obj.cell(row=i, column=j)
                v.append(v1.value)
            d[key] = v

        value =Student.objects.filter(Branch=branch,Sem=sem,Div=div).count()
        if len(d) != value:
            queryset = Student.objects.filter(Branch=branch,Sem=sem,Div=div)
            for student in queryset:
                name=student.Name
                roll=student.Roll
                if not roll in d:
                    # l = []
                    d[roll] = []
                    d[roll].append(name)
                    print(f"if {roll} {name} {d[roll]}")
            r = 3

            print(d)
            for key, value in d.items():
                col = 1
                print(key, end=' ')
                sheet_obj.cell(row=r, column=col).value = key

                for list in value:
                    col += 1
                    sheet_obj.cell(row=r, column=col).value = list
                    print(list, end=' ')
                r += 1

        for col_index in range(1, 100):
            col = get_column_letter(col_index)
            if sheet.cell('%s%s' % (col, 1)).value is None:
                col2 = get_column_letter(col_index - 1)
                # print sheet.cell('%s%s'% (col2, 1)).value
                if sheet.cell('%s%s' % (col2, 1)).value != currentDate:
                    sheet['%s%s' % (col, 1)] = currentDate
                break

        # saving the file
        wb.save(filename=reportName)

    else:
        wb = Workbook()
        dest_filename = reportName
        sortedData=Student.objects.filter(Branch=branch,Sem=sem,Div=div)
        # creating worksheet and giving names to column
        ws1 = wb.active
        ws1.title = "Cse15"
        ws1.append(('Roll Number', 'Name', currentDate))
        ws1.append(('', '', ''))

        # entering students information from database
        for student in sortedData:
            ws1.append((student.Roll,student.Name))

        # saving the file
        wb.save(filename=dest_filename)
def identify(emailt,subject,branch,sem,div):
    currentDate = time.strftime("%d_%m_%y")
    reportName = emailt + "_" + subject + "_" + branch + "_" + sem + "_" + div + '.xlsx'
    wb = load_workbook(filename=reportName)
    sheet = wb.get_sheet_by_name('Cse15')

    def getDateColumn():
        for i in range(1, len(sheet.rows[0]) + 1):
            col = get_column_letter(i)
            if sheet.cell('%s%s' % (col, '1')).value == currentDate:
                return col

    Key = '97849dde33244e88a570f60327c27604'
    CF.Key.set(Key)

    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)
    # connect = connect = sqlite3.connect("Face-DataBase")
    # c = connect.cursor()

    attend = [0 for i in range(150)]

    currentDir = os.path.dirname(os.path.abspath(__file__))
    directory = os.path.join(currentDir, 'Cropped_faces')
    print(personGroupId())
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = os.path.join(directory, filename)

            res = CF.face.detect(imgurl)
            print(imgurl)
            print(res)
            if len(res) != 1:
                print("No face detected.")
                continue

            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            print(f'{faceIds} {personGroupId()}')

            res = CF.face.identify(faceIds, personGroupId())
            print(filename)
            print(res)
            for face in res:
                if not face['candidates']:
                    print("Unknown")
                else:
                    personId = face['candidates'][0]['personId']
                    # c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                    # row = c.fetchone()
                    row = getDataUsingPersonId(personId,branch,sem,div)
                    if row is None:
                        continue
                    id = int(row.Roll[-2:])
                    attend[id] += 1
                    print(row.Name + " recognized")
            time.sleep(6)
    # print(len(sheet.columns[0]))
    # for r in range(2, len(sheet.columns[0]) + 1):
    for i in range(2,sheet.max_row+1):
        roll = sheet.cell('A%s' % i).value
        if roll is not None:
            id = int(roll[-2:])
            if attend[id] != 0:
                col = getDateColumn()
                sheet['%s%s' % (col, str(i))] = 1

    wb.save(filename=reportName)
    # return redirect('success')

def success(request):
    return HttpResponse('successfully done')

def error(request):
    return HttpResponse('something might wrong')
def generateAttendance(request):
    if request.method == 'POST':
        form = GenerateAttendance(request.POST)
        if form.is_valid():
            subject=form.cleaned_data.get("subject")
            branch = form.cleaned_data.get("branch")
            sem = form.cleaned_data.get("sem")
            div = form.cleaned_data.get("div")
            response = HttpResponse(content_type='application/ms-excel')
            reportName=request.session['emailt']+"_"+ subject + "_" + branch + "_" + sem + "_" + div + ".xlsx"
            response['Content-Disposition'] = 'attachment; filename="'+reportName+'"'

            wb1 = xlwt.Workbook(encoding='utf-8')
            ws = wb1.add_sheet('cse20')  # this will make a sheet named Users Data

            if (os.path.exists('./' + reportName)):
                wb = load_workbook(filename=reportName)
                sheet = wb.get_sheet_by_name('Cse15')
                # sheet[ord() + '1']

                sheet_obj = wb.active
                m_row = sheet_obj.max_row
                m_column = sheet_obj.max_column
                list=[]
                for i in range(1, m_row + 1):
                    # key = sheet_obj.cell(row=i, column=1).value
                    v = []
                    for j in range(1, m_column + 1):
                        v1 = sheet_obj.cell(row=i, column=j)
                        v.append(v1.value)
                    list.append(v)

                row_num = 0

                font_style = xlwt.XFStyle()
                # font_style.font.bold = True

                # columns = ['Enrollment', 'First Name', 'Last Name', 'Email Address', ]

                # for col_num in range(len(columns)):
                #   ws.write(row_num, col_num, columns[col_num], font_style)  # at 0 row 0 column

                # Sheet body, remaining rows
                font_style = xlwt.XFStyle()
                for row_index, row in enumerate(list):
                    for col_index, cell_value in enumerate(row):
                        ws.write(row_index, col_index, cell_value)


                wb1.save(response)

            else:
                print("sheet is not found")
            # Sheet header, first row


            return response
    else:
        form = GenerateAttendance()
        print("error")
    return render(request, 'generateattendance.html', {'form': form})
